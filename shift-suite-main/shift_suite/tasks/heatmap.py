# shift_suite / tasks / heatmap.py
# v1.8.1 (日曜日Need計算修正版)
from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import List, Set

import numpy as np
import openpyxl
import pandas as pd
import logging
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .constants import SUMMARY5
from shift_suite.i18n import translate as _

# 'log' という名前でロガーを取得 (utils.pyからインポートされるlogと同じ)
from .utils import (
    _parse_as_date,
    derive_max_staff,
    gen_labels,
    log,
    safe_sheet,
    save_df_xlsx,
    write_meta,
    validate_need_calculation,
)

analysis_logger = logging.getLogger('analysis')

# 新規追加: 通常勤務の判定用定数
DEFAULT_HOLIDAY_TYPE = "通常勤務"

STAFF_ALIASES = ["staff", "氏名", "名前", "従業員"]
ROLE_ALIASES = ["role", "職種", "役職", "部署"]


def _resolve(df: pd.DataFrame, prefer: str, aliases: list[str], new: str) -> str:
    if prefer in df.columns:
        df.rename(columns={prefer: new}, inplace=True)
        return new
    for alias_name in aliases:
        if alias_name in df.columns:
            df.rename(columns={alias_name: new}, inplace=True)
            return new
    raise KeyError(
        f"列 '{prefer}' またはそのエイリアス {aliases} がDataFrameに見つかりません。"
    )


def _apply_conditional_formatting_to_worksheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, df_data_columns: pd.Index
):
    log.debug(f"[heatmap._apply_cf] 書式設定対象のワークシート: '{worksheet.title}'")
    log.debug(
        f"[heatmap._apply_cf] 書式設定の基準となるデータ列 (df_data_columns): {df_data_columns.tolist() if isinstance(df_data_columns, pd.Index) else df_data_columns}"
    )
    if worksheet.max_row <= 1:
        return
    if (
        df_data_columns.empty
        if isinstance(df_data_columns, pd.Index)
        else not df_data_columns
    ):
        return
    first_data_col_excel_idx = 2
    last_data_col_excel_idx = (
        first_data_col_excel_idx + (len(df_data_columns) - 1)
        if (isinstance(df_data_columns, pd.Index) and not df_data_columns.empty)
        or (not isinstance(df_data_columns, pd.Index) and df_data_columns)
        else first_data_col_excel_idx - 1
    )
    if last_data_col_excel_idx < first_data_col_excel_idx:
        return
    range_start_cell = "B2"
    range_end_col_letter = get_column_letter(last_data_col_excel_idx)
    range_end_row_num = worksheet.max_row
    data_range_string = f"{range_start_cell}:{range_end_col_letter}{range_end_row_num}"
    log.info(
        f"[heatmap._apply_cf] シート '{worksheet.title}' に条件付き書式を適用します。範囲: {data_range_string}"
    )
    try:
        color_scale_rule = ColorScaleRule(
            start_type="min",
            start_color="FFFFE0",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFA500",
            end_type="max",
            end_color="FF0000",
        )
        worksheet.conditional_formatting.add(data_range_string, color_scale_rule)
    except Exception as e:
        log.error(f"[heatmap._apply_cf] 条件付き書式適用中にエラー: {e}", exc_info=True)


def _apply_holiday_column_styling(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    date_columns_in_excel: pd.Index,
    estimated_holidays: Set[dt.date],
    utils_parse_as_date_func,
):
    if not estimated_holidays or date_columns_in_excel.empty:
        return
    holiday_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    first_data_col_excel_letter_idx = 2
    for i, col_name_excel_str in enumerate(date_columns_in_excel):
        current_col_date = utils_parse_as_date_func(str(col_name_excel_str))
        if current_col_date and current_col_date in estimated_holidays:
            target_excel_col_idx = first_data_col_excel_letter_idx + i
            col_letter = get_column_letter(target_excel_col_idx)
            for row_idx in range(1, worksheet.max_row + 1):
                worksheet[f"{col_letter}{row_idx}"].fill = holiday_fill


def calculate_pattern_based_need(
    actual_staff_by_slot_and_date: pd.DataFrame,
    ref_start_date: dt.date,
    ref_end_date: dt.date,
    statistic_method: str,
    remove_outliers: bool,
    iqr_multiplier: float = 1.5,
    slot_minutes_for_empty: int = 30,
    *,
    holidays: set[dt.date] | None = None,
    adjustment_factor: float = 1.0,
    include_zero_days: bool = True,
    all_dates_in_period: list[dt.date] | None = None,
) -> pd.DataFrame:
    # 修正箇所: logger.info -> log.info など、ロガー名を 'log' に統一
    log.info(
        f"[heatmap.calculate_pattern_based_need] 参照期間: {ref_start_date} - {ref_end_date}, 手法: {statistic_method}, 外れ値除去: {remove_outliers}"
    )

    time_index_labels = pd.Index(gen_labels(slot_minutes_for_empty), name="time")
    default_dow_need_df = pd.DataFrame(
        0, index=time_index_labels, columns=range(7)
    )  # 月曜0 - 日曜6

    if actual_staff_by_slot_and_date.empty:
        log.warning(
            "[heatmap.calculate_pattern_based_need] 入力実績データが空です。デフォルトの0 needを返します。"
        )
        return default_dow_need_df

    # actual_staff_by_slot_and_date の列名が日付オブジェクトであることを確認・変換
    # 呼び出し元(build_heatmap)で列名をdt.dateオブジェクトに変換済みのものを渡すように修正
    df_for_calc = actual_staff_by_slot_and_date.copy()

    holidays_set = set(holidays or [])

    if include_zero_days:
        log.info("[NEED_FIX] include_zero_days=True → 休業日除外なし")

    # 重要な修正：全期間の日付を考慮する
    if all_dates_in_period and include_zero_days:
        all_dates_in_ref = [
            d for d in all_dates_in_period
            if isinstance(d, dt.date) and ref_start_date <= d <= ref_end_date and d not in holidays_set
        ]

        # 実績がない日付を0で埋める
        for date in all_dates_in_ref:
            if date not in df_for_calc.columns:
                df_for_calc[date] = 0

        log.info(f"[NEED_FIX] 全期間の日付を考慮: 元の列数={len(actual_staff_by_slot_and_date.columns)}, 補完後={len(df_for_calc.columns)}")

    # 参照期間でデータをフィルタリング (列名がdt.dateオブジェクトであることを前提)
    cols_to_process_dow = [
        col_date
        for col_date in df_for_calc.columns
        if (
            isinstance(col_date, dt.date)
            and ref_start_date <= col_date <= ref_end_date
            and col_date not in holidays_set
        )
    ]

    if not cols_to_process_dow:
        log.warning(
            f"[heatmap.calculate_pattern_based_need] 参照期間 ({ref_start_date} - {ref_end_date}) に該当する実績データがありません。"
        )
        return default_dow_need_df

    filtered_slot_df_dow = df_for_calc[cols_to_process_dow]
    dow_need_df_calculated = pd.DataFrame(
        index=filtered_slot_df_dow.index, columns=range(7), dtype=float
    )

    for day_of_week_idx in range(7):
        dow_cols_to_agg = [
            col_dt
            for col_dt in filtered_slot_df_dow.columns
            if col_dt.weekday() == day_of_week_idx
        ]

        # デバッグ: 曜日名マッピング
        dow_names = {0: "月曜日", 1: "火曜日", 2: "水曜日", 3: "木曜日", 4: "金曜日", 5: "土曜日", 6: "日曜日"}
        dow_name = dow_names.get(day_of_week_idx, f"曜日{day_of_week_idx}")
        log.info(f"[NEED_DEBUG] === {dow_name} ({day_of_week_idx}) 処理開始 ===")
        log.info(f"[NEED_DEBUG] 対象日付数: {len(dow_cols_to_agg)}")

        if not dow_cols_to_agg:
            log.warning(f"[NEED_DEBUG] {dow_name}: 対象データなし")
            dow_need_df_calculated[day_of_week_idx] = 0
            continue

        # デバッグ情報出力（全曜日）
        log.info(
            f"[NEED_DEBUG] 対象日付例: {[d.strftime('%Y-%m-%d') for d in dow_cols_to_agg[:3]]}{'...' if len(dow_cols_to_agg) > 3 else ''}"
        )

        data_for_dow_calc = filtered_slot_df_dow[dow_cols_to_agg]

        is_significant_holiday = False
        if not data_for_dow_calc.empty:
            avg_staff_per_day_overall = filtered_slot_df_dow.sum().mean()
            avg_staff_per_day_dow = data_for_dow_calc.sum().mean()
            analysis_logger.info(
                f"曜日 '{dow_name}'({day_of_week_idx}) の必要人数計算: "
                f"曜日別平均勤務人数 = {avg_staff_per_day_dow:.2f}, "
                f"全体平均勤務人数 = {avg_staff_per_day_overall:.2f}, "
                f"適用中の統計手法 = '{statistic_method}'"
            )
            # 日曜日は強制的に特殊処理対象とする
            if day_of_week_idx == 6:  # 日曜日
                is_significant_holiday = True
                analysis_logger.info("[SUNDAY_FORCE] 日曜日のため強制的に特殊処理を適用します")
            elif avg_staff_per_day_dow < (avg_staff_per_day_overall * 0.25):
                analysis_logger.warning(
                    f"曜日 '{dow_name}'({day_of_week_idx}) は勤務実績が著しく少ないため、"
                    f"必要人数が実態と乖離する可能性があります。"
                )
                is_significant_holiday = True

        # 日毎の合計人数を計算
        daily_totals = data_for_dow_calc.sum()
        log.info(f"  各日の総勤務人数: {daily_totals.values.tolist()}")
        log.info(f"  日平均総勤務人数: {daily_totals.mean():.2f}")

        # 時間帯別の詳細（特に日曜日は詳細に）
        if day_of_week_idx == 6:
            log.info("[SUNDAY_DEBUG] ========== 日曜日の詳細分析 ==========")
            log.info("[SUNDAY_DEBUG] 対象期間の全日曜日:")
            for d in dow_cols_to_agg:
                daily_sum = data_for_dow_calc[d].sum()
                log.info(f"[SUNDAY_DEBUG]   {d.strftime('%Y-%m-%d')}: {daily_sum}名")

            log.info("[SUNDAY_DEBUG] 代表的な時間帯の値:")
            sample_times = ["09:00", "12:00", "15:00", "18:00"]
            for time_slot in sample_times:
                if time_slot in data_for_dow_calc.index:
                    values = data_for_dow_calc.loc[time_slot].values.tolist()
                    log.info(f"[SUNDAY_DEBUG]   {time_slot}: {values}")
        # ▼▼▼ ロジック修正 ▼▼▼
        # 統計手法を決定する
        # 実データが少ない場合の統計手法調整
        if is_significant_holiday:
            current_statistic_method = "中央値"
            analysis_logger.info(
                f" -> 曜日 '{dow_name}' は実績僅少のため、統計手法を「{current_statistic_method}」に自動調整しました。"
            )
        else:
            current_statistic_method = statistic_method

        for time_slot_val, row_series_data in data_for_dow_calc.iterrows():
            if include_zero_days:
                values_at_slot_current = [0.0 if pd.isna(v) else float(v) for v in row_series_data]
            else:
                values_at_slot_current = row_series_data.dropna().astype(float).tolist()
            analysis_logger.info(
                f"[DEBUG_NEED_DETAIL] 処理中の時間帯: {time_slot_val} ({dow_name}), 元データ ({len(values_at_slot_current)}点): {values_at_slot_current}"
            )

            if not values_at_slot_current:
                dow_need_df_calculated.loc[time_slot_val, day_of_week_idx] = 0
                continue
            values_for_stat_calc = values_at_slot_current
            if day_of_week_idx == 6 and time_slot_val in ["09:00", "12:00", "15:00"]:
                log.info(f"[SUNDAY_DETAIL] {time_slot_val} 時間帯:")
                log.info(f"[SUNDAY_DETAIL]   元データ: {values_at_slot_current}")
            # 統計値の計算前にデバッグ情報を出力
            if day_of_week_idx == 6 or (day_of_week_idx == 1 and time_slot_val == "09:00"):
                log.info(f"\n  [統計計算デバッグ] {dow_name} {time_slot_val}")
                log.info(f"    元データ: {values_at_slot_current}")
                log.info(f"    データ数: {len(values_at_slot_current)}")

            if remove_outliers and len(values_at_slot_current) >= 4:
                q1_val = np.percentile(values_at_slot_current, 25)
                q3_val = np.percentile(values_at_slot_current, 75)
                iqr_val = q3_val - q1_val
                lower_bound_val = q1_val - iqr_multiplier * iqr_val
                upper_bound_val = q3_val + iqr_multiplier * iqr_val
                values_filtered_outlier = [
                    x_val
                    for x_val in values_at_slot_current
                    if lower_bound_val <= x_val <= upper_bound_val
                ]
                # デバッグ: 外れ値除去の詳細
                if day_of_week_idx == 6 and time_slot_val in ["09:00", "12:00", "15:00"]:
                    log.info(f"[SUNDAY_DETAIL]   外れ値除去後: {values_filtered_outlier}")

                analysis_logger.info(
                    f"[DEBUG_NEED_DETAIL] 外れ値除去実行前 (Q1:{q1_val:.1f}, Q3:{q3_val:.1f}, IQR:{iqr_val:.1f}), フィルタリング後 ({len(values_filtered_outlier)}点): {values_filtered_outlier}"
                )

                if not values_filtered_outlier:
                    log.debug(
                        f"  曜日 {day_of_week_idx}, 時間帯 {time_slot_val}: 外れ値除去後データなし。元のリストで計算します。"
                    )
                else:
                    values_for_stat_calc = values_filtered_outlier
            need_calculated_val = 0.0
            if values_for_stat_calc:
                # 決定された統計手法に基づいて計算
                if current_statistic_method == "10パーセンタイル":
                    need_calculated_val = np.percentile(values_for_stat_calc, 10)
                elif current_statistic_method == "25パーセンタイル":
                    need_calculated_val = np.percentile(values_for_stat_calc, 25)
                elif current_statistic_method == "中央値":
                    need_calculated_val = np.median(values_for_stat_calc)
                elif current_statistic_method == "75パーセンタイル":
                    need_calculated_val = np.percentile(values_for_stat_calc, 75)
                elif current_statistic_method == "90パーセンタイル":
                    need_calculated_val = np.percentile(values_for_stat_calc, 90)
                else:  # 平均値
                    need_calculated_val = np.mean(values_for_stat_calc)
            analysis_logger.info(
                f"[DEBUG_NEED_DETAIL] 統計手法({current_statistic_method})適用後のNeed仮値: {need_calculated_val:.2f}"
            )

            # データの中央値が小さい場合はNeedを上限2.0に制限
            if values_at_slot_current and np.median(values_at_slot_current) < 2.0:
                need_calculated_val = min(need_calculated_val, 2.0)
                analysis_logger.info(
                    f"  [NEED_CAP] 曜日 {day_of_week_idx}, 時間帯 {time_slot_val}: "
                    f"実績中央値が2未満のためNeedを {need_calculated_val:.1f} に制限しました。"
                )
                analysis_logger.info(
                    f"[DEBUG_NEED_DETAIL] Need上限適用判定: 元データ中央値={np.median(values_at_slot_current):.1f}。制限後Need={need_calculated_val:.2f}"
                )

            # 調整係数の適用
            need_calculated_val *= adjustment_factor
            
            # 実データが少ない場合の特殊処理
            if is_significant_holiday:
                # データが少ない場合は、実際の最大値を上限として設定
                max_actual_val = max(values_at_slot_current) if values_at_slot_current else 0
                if need_calculated_val > max_actual_val * 1.5:  # 実際の最大値の1.5倍を上限
                    original_need = need_calculated_val
                    need_calculated_val = max_actual_val * 1.5
                    log.info(f"[STATS_FIX] {dow_name} {time_slot_val}: Need値を {original_need:.2f} → {need_calculated_val:.2f} に制限（実データ考慮）")
                
                # さらに、0が多いデータでは0により近い値に調整
                zero_ratio = values_at_slot_current.count(0) / len(values_at_slot_current) if values_at_slot_current else 1
                if zero_ratio > 0.5:  # 50%以上が0の場合
                    need_calculated_val *= (1 - zero_ratio * 0.5)  # 0の比率に応じて減算
                    log.info(f"[STATS_FIX] {dow_name} {time_slot_val}: 0データ比率{zero_ratio:.2f}により調整 → {need_calculated_val:.2f}")
            
            final_need = round(need_calculated_val) if not pd.isna(need_calculated_val) else 0
            dow_need_df_calculated.loc[time_slot_val, day_of_week_idx] = final_need
            log.debug(
                f"  曜日 {day_of_week_idx}, 時間帯 {time_slot_val}: 元データ長 {len(row_series_data.dropna())} -> 外れ値除去後 {len(values_for_stat_calc)} -> Need {dow_need_df_calculated.loc[time_slot_val, day_of_week_idx]}"
            )

    # 全曜日の計算完了後、サマリーを出力
    log.info("[NEED_DEBUG] ========== Need計算完了サマリー ==========")
    for dow_idx in range(7):
        dow_name = dow_names.get(dow_idx, f"曜日{dow_idx}")
        total_need = dow_need_df_calculated[dow_idx].sum()
        max_need = dow_need_df_calculated[dow_idx].max()
        avg_need = dow_need_df_calculated[dow_idx].mean()
        log.info(f"[NEED_DEBUG] {dow_name}: 合計={total_need:.0f}, 最大={max_need:.0f}, 平均={avg_need:.2f}")

    # 特に日曜日の詳細
    if 6 in dow_need_df_calculated.columns:
        sunday_data = dow_need_df_calculated[6]
        log.info("\n[日曜日の時間帯別Need値]")
        for time_slot, need_val in sunday_data.items():
            if need_val > 0:
                log.info(f"  {time_slot}: {need_val}")

    log.info("[heatmap.calculate_pattern_based_need] 曜日別・時間帯別needの算出完了。")
    return dow_need_df_calculated.fillna(0).astype(int)


def _filter_work_records(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    新規追加: 通常勤務のレコードのみを抽出する
    休暇レコード（holiday_type != "通常勤務"）を除外し、
    実際に勤務時間がある（parsed_slots_count > 0）レコードのみを返す
    """
    if long_df.empty:
        return long_df

    # 通常勤務且つ勤務時間があるレコードのみ抽出
    work_records = long_df[
        (long_df.get("holiday_type", DEFAULT_HOLIDAY_TYPE) == DEFAULT_HOLIDAY_TYPE)
        & (long_df.get("parsed_slots_count", 0) > 0)
    ].copy()

    original_count = len(long_df)
    work_count = len(work_records)
    leave_count = original_count - work_count

    log.info(
        f"[heatmap._filter_work_records] フィルタリング結果: 全レコード={original_count}, 勤務レコード={work_count}, 休暇レコード={leave_count}"
    )

    if not work_records.empty:
        holiday_stats = long_df["holiday_type"].value_counts()
        log.debug(f"[heatmap._filter_work_records] 休暇タイプ別統計:\n{holiday_stats}")

    return work_records


def build_heatmap(
    long_df: pd.DataFrame,
    out_dir: str | Path,
    slot_minutes: int = 30,
    *,
    need_calc_method: str | None = None,
    need_stat_method: str | None = None,
    include_zero_days: bool = True,
    need_manual_values: dict | None = None,
    upper_calc_method: str | None = None,
    upper_calc_param: dict | None = None,
    # legacy parameters kept for backward compatibility
    ref_start_date_for_need: dt.date | None = None,
    ref_end_date_for_need: dt.date | None = None,
    need_statistic_method: str | None = None,
    need_remove_outliers: bool | None = None,
    need_iqr_multiplier: float | None = 1.5,
    need_adjustment_factor: float = 1.0,
    min_method: str = "p25",
    max_method: str = "p75",
    holidays: set[dt.date] | None = None,
) -> None:
    holidays_set = set(holidays or [])

    if long_df.empty:
        log.warning("[heatmap.build_heatmap] 入力DataFrame (long_df) が空です。")
        return
    required_long_df_cols = {
        "ds",
        "staff",
        "role",
        "code",
        "holiday_type",
        "parsed_slots_count",
    }
    if not required_long_df_cols.issubset(long_df.columns):
        missing_cols = required_long_df_cols - set(long_df.columns)
        log.error(f"[heatmap.build_heatmap] long_dfに必要な列 {missing_cols} が不足。")
        out_dir_path = Path(out_dir)
        out_dir_path.mkdir(parents=True, exist_ok=True)
        write_meta(
            out_dir_path / "heatmap.meta.json",
            slot=slot_minutes,
            roles=[],
            dates=[],
            summary_columns=SUMMARY5,
            estimated_holidays=[d.isoformat() for d in sorted(list(holidays or set()))],
        )
        return

    # 重要: 休暇レコードの統計を先に収集
    leave_stats = {}
    if not long_df.empty and "holiday_type" in long_df.columns:
        holiday_type_stats = long_df["holiday_type"].value_counts()
        leave_stats = {
            "total_records": len(long_df),
            "leave_records": len(
                long_df[long_df["holiday_type"] != DEFAULT_HOLIDAY_TYPE]
            ),
            "holiday_type_breakdown": holiday_type_stats.to_dict(),
        }
        log.info(f"[heatmap.build_heatmap] 休暇統計: {leave_stats}")
    estimated_holidays_set: Set[dt.date] = set()
    all_dates_in_period_list: List[dt.date] = []
    if (
        not long_df.empty
        and "ds" in long_df.columns
        and "parsed_slots_count" in long_df.columns
    ):
        long_df_for_holiday_check = long_df.copy()
        if not pd.api.types.is_datetime64_any_dtype(long_df_for_holiday_check["ds"]):
            long_df_for_holiday_check["ds"] = pd.to_datetime(
                long_df_for_holiday_check["ds"], errors="coerce"
            )
        valid_ds_long_df = long_df_for_holiday_check.dropna(subset=["ds"])
        if not valid_ds_long_df.empty:
            min_date_val = valid_ds_long_df["ds"].dt.date.min()
            max_date_val = valid_ds_long_df["ds"].dt.date.max()
            if (
                pd.NaT not in [min_date_val, max_date_val]
                and isinstance(min_date_val, dt.date)
                and isinstance(max_date_val, dt.date)
            ):
                current_scan_date = min_date_val
                while current_scan_date <= max_date_val:
                    all_dates_in_period_list.append(current_scan_date)
                    current_scan_date += dt.timedelta(days=1)
            else:
                log.warning("[heatmap.build_heatmap] 有効な日付範囲を決定できません。")
            if all_dates_in_period_list:
                for current_date_val_iter in all_dates_in_period_list:
                    df_for_current_date_iter = valid_ds_long_df[
                        valid_ds_long_df["ds"].dt.date == current_date_val_iter
                    ]
                    if df_for_current_date_iter.empty:
                        estimated_holidays_set.add(current_date_val_iter)
                        log.debug(
                            f"施設休業日(推定): {current_date_val_iter} (勤務記録なし)"
                        )
                        continue

                    # 修正: 通常勤務のレコードのみで判定
                    work_records_today = _filter_work_records(df_for_current_date_iter)
                    if work_records_today.empty:
                        estimated_holidays_set.add(current_date_val_iter)
                        log.debug(
                            f"施設休業日(推定): {current_date_val_iter} (通常勤務なし)"
                        )
            if estimated_holidays_set:
                log.info(
                    f"[heatmap.build_heatmap] 推定された休業日 ({len(estimated_holidays_set)}日): {sorted(list(estimated_holidays_set))}"
                )
            else:
                log.info("[heatmap.build_heatmap] 推定される休業日はありませんでした。")
        else:
            log.warning(
                "[heatmap.build_heatmap] 'ds'列に有効な日時データがないため、休業日を推定できません。"
            )
    else:
        log.warning(
            "[heatmap.build_heatmap] long_dfが空か、必要な列がないため、休業日を推定できません。"
        )

    # 重要: 通常勤務のレコードのみでヒートマップ作成
    df_for_heatmap_actuals = _filter_work_records(long_df)
    out_dir_path = Path(out_dir)
    out_dir_path.mkdir(parents=True, exist_ok=True)
    all_date_labels_in_period_str: List[str] = (
        sorted([d.strftime("%Y-%m-%d") for d in all_dates_in_period_list])
        if all_dates_in_period_list
        else []
    )
    time_index_labels = pd.Index(gen_labels(slot_minutes), name="time")

    if df_for_heatmap_actuals.empty and not all_date_labels_in_period_str:
        log.warning(
            "[heatmap.build_heatmap] 有効な勤務データも日付範囲もないため、ヒートマップは空になります。"
        )
        empty_pivot = pd.DataFrame(index=time_index_labels)
        for col_name_ep_loop in SUMMARY5:
            empty_pivot[col_name_ep_loop] = 0
        fp_all_empty_path = out_dir_path / "heat_ALL.parquet"
        try:
            empty_pivot.to_parquet(fp_all_empty_path)
        except Exception as e_empty_write:
            log.error(f"空のheat_ALL.parquetの書き込みに失敗: {e_empty_write}")
        all_unique_roles_val = (
            sorted(list(set(long_df["role"]))) if "role" in long_df.columns else []
        )
        write_meta(
            out_dir_path / "heatmap.meta.json",
            slot=slot_minutes,
            roles=all_unique_roles_val,
            dates=[],
            summary_columns=SUMMARY5,
            estimated_holidays=[d.isoformat() for d in sorted(list(holidays or set()))],
            leave_statistics=leave_stats,  # 休暇統計を追加
        )
        return

    df_for_heatmap_actuals["time"] = pd.to_datetime(
        df_for_heatmap_actuals["ds"], errors="coerce"
    ).dt.strftime("%H:%M")
    df_for_heatmap_actuals["date_lbl"] = pd.to_datetime(
        df_for_heatmap_actuals["ds"], errors="coerce"
    ).dt.strftime("%Y-%m-%d")
    df_for_heatmap_actuals.dropna(
        subset=["time", "date_lbl", "staff", "role"], inplace=True
    )

    staff_col_name = "staff"
    role_col_name = "role"
    log.info("[heatmap.build_heatmap] 全体ヒートマップ作成開始。")

    pivot_data_all_actual_staff = pd.DataFrame(index=time_index_labels)
    if not df_for_heatmap_actuals.empty:
        if len(df_for_heatmap_actuals) > 50000:
            log.info(
                "[heatmap.build_heatmap] Large dataset detected, using chunked processing"
            )
            chunk_size = 10000
            pivot_chunks = []

            for i in range(0, len(df_for_heatmap_actuals), chunk_size):
                chunk = df_for_heatmap_actuals.iloc[i : i + chunk_size]
                chunk_pivot = chunk.drop_duplicates(
                    subset=["date_lbl", "time", staff_col_name]
                ).pivot_table(
                    index="time",
                    columns="date_lbl",
                    values=staff_col_name,
                    aggfunc="nunique",
                    fill_value=0,
                )
                pivot_chunks.append(chunk_pivot)

            if pivot_chunks:
                pivot_data_all_actual_staff = pd.concat(pivot_chunks, axis=1)
                pivot_data_all_actual_staff = pivot_data_all_actual_staff.groupby(
                    level=0, axis=1
                ).sum()
                pivot_data_all_actual_staff = pivot_data_all_actual_staff.reindex(
                    index=time_index_labels, fill_value=0
                )
        else:
            pivot_data_all_actual_staff = (
                df_for_heatmap_actuals.drop_duplicates(
                    subset=["date_lbl", "time", staff_col_name]
                )
                .pivot_table(
                    index="time",
                    columns="date_lbl",
                    values=staff_col_name,
                    aggfunc="nunique",
                    fill_value=0,
                )
                .reindex(index=time_index_labels, fill_value=0)
            )

    # Ensure all dates in the period are present as columns, filling missing ones with 0
    pivot_data_all_actual_staff = pivot_data_all_actual_staff.reindex(
        columns=all_date_labels_in_period_str, fill_value=0
    )

    actual_staff_for_need_input = pivot_data_all_actual_staff.copy()
    if not actual_staff_for_need_input.empty:
        new_column_map_for_need_input = {}
        for col_str_need in actual_staff_for_need_input.columns:
            dt_obj_need = _parse_as_date(str(col_str_need))
            if dt_obj_need:
                # parsed_date_columns_for_need_input.append(dt_obj_need) # これは不要
                new_column_map_for_need_input[col_str_need] = dt_obj_need
            else:
                log.debug(
                    f"Need計算用実績データの列名'{col_str_need}'を日付にパースできませんでした。"
                )
        if new_column_map_for_need_input:
            # renameする前に、キー(元の列名)が存在するか確認
            valid_keys_for_rename = [
                k
                for k in new_column_map_for_need_input.keys()
                if k in actual_staff_for_need_input.columns
            ]
            actual_staff_for_need_input = actual_staff_for_need_input[
                valid_keys_for_rename
            ].rename(columns=new_column_map_for_need_input)
        else:
            actual_staff_for_need_input = pd.DataFrame(index=time_index_labels)

    # app.pyから渡される新しい引数(need_stat_method)を優先し、
    # legacy引数(need_statistic_method)があればバックアップとして使用する
    final_statistic_method = (
        need_stat_method if need_stat_method is not None else need_statistic_method
    )

    if include_zero_days:
        log.info("[NEED_FIX] include_zero_days=True → 推定休業日を無視")
        final_holidays_to_use = holidays_set
    else:
        final_holidays_to_use = holidays_set.union(estimated_holidays_set)

    overall_dow_need_pattern_df = calculate_pattern_based_need(
        actual_staff_for_need_input,
        ref_start_date_for_need,
        ref_end_date_for_need,
        final_statistic_method,
        need_remove_outliers,
        need_iqr_multiplier,
        slot_minutes_for_empty=slot_minutes,
        holidays=final_holidays_to_use,
        adjustment_factor=need_adjustment_factor,
        include_zero_days=include_zero_days,
        all_dates_in_period=all_dates_in_period_list,
    )

    pivot_data_all_final = pd.DataFrame(
        index=time_index_labels, columns=all_date_labels_in_period_str, dtype=float
    ).fillna(0)
    need_all_final_for_summary = pd.DataFrame(
        index=time_index_labels, columns=all_date_labels_in_period_str, dtype=float
    ).fillna(0)

    for date_str_col_map in all_date_labels_in_period_str:
        if date_str_col_map in pivot_data_all_actual_staff.columns:
            pivot_data_all_final[date_str_col_map] = pivot_data_all_actual_staff[
                date_str_col_map
            ]
        current_date_obj_map = dt.datetime.strptime(date_str_col_map, "%Y-%m-%d").date()
        if current_date_obj_map.weekday() == 6:
            log.info(f"[SUNDAY_APPLY] 日曜日 {date_str_col_map} のNeed適用:")
            log.info(f"[SUNDAY_APPLY]   休業日判定: {current_date_obj_map in holidays_set}")
        if current_date_obj_map in holidays_set:
            if current_date_obj_map.weekday() == 6:
                log.info("[SUNDAY_APPLY]   → 休業日のためNeed=0に設定")
            need_all_final_for_summary[date_str_col_map] = 0
        else:
            day_of_week_map = current_date_obj_map.weekday()
            if day_of_week_map in overall_dow_need_pattern_df.columns:
                need_all_final_for_summary[date_str_col_map] = overall_dow_need_pattern_df[day_of_week_map]
                if current_date_obj_map.weekday() == 6:
                    need_values = overall_dow_need_pattern_df[day_of_week_map]
                    log.info(f"[SUNDAY_APPLY]   → Need値適用: 合計={need_values.sum():.0f}")
                    log.info(f"[SUNDAY_APPLY]   → Need値詳細（最初5つ）: {need_values.head().tolist()}")

            else:
                need_all_final_for_summary[date_str_col_map] = 0
                log.warning(
                    f"曜日 {day_of_week_map} のneedパターンが見つかりません ({date_str_col_map})。Needは0とします。"
                )

    # 詳細なNeedデータをParquetファイルとして保存
    need_all_final_for_summary.to_parquet(
        out_dir_path / "need_per_date_slot.parquet"
    )
    log.info("Need per date/slot data saved to need_per_date_slot.parquet.")

    upper_s_representative = (
        derive_max_staff(pivot_data_all_actual_staff, max_method)
        if not pivot_data_all_actual_staff.empty
        else pd.Series(0, index=time_index_labels)
    )

    total_lack_per_time = pd.Series(0, index=time_index_labels, dtype=float)
    total_excess_per_time = pd.Series(0, index=time_index_labels, dtype=float)
    working_day_count = 0

    for date_col in need_all_final_for_summary.columns:
        if date_col in pivot_data_all_final.columns:
            daily_staff = pivot_data_all_final[date_col]
            daily_need = need_all_final_for_summary[date_col]

            date_obj = dt.datetime.strptime(date_col, "%Y-%m-%d").date()
            if date_obj not in holidays_set:
                working_day_count += 1
                daily_lack = (daily_need - daily_staff).clip(lower=0)
                daily_excess = (daily_staff - upper_s_representative).clip(lower=0)

                total_lack_per_time += daily_lack
                total_excess_per_time += daily_excess

    avg_need_series = need_all_final_for_summary.mean(axis=1).round()
    avg_staff_series = (
        pivot_data_all_final.drop(columns=SUMMARY5, errors="ignore")
        .mean(axis=1)
        .round()
        if not pivot_data_all_final.empty
        else pd.Series(0, index=time_index_labels)
    )
    avg_lack_series = (total_lack_per_time / max(working_day_count, 1)).round()
    avg_excess_series = (total_excess_per_time / max(working_day_count, 1)).round()

    pivot_to_excel_all = pivot_data_all_final.copy()
    for col_name_summary_loop, series_data_summary_loop in zip(
        SUMMARY5,
        [
            avg_need_series,
            upper_s_representative,
            avg_staff_series,
            avg_lack_series,
            avg_excess_series,
        ],
        strict=True,
    ):
        pivot_to_excel_all[col_name_summary_loop] = series_data_summary_loop

    analysis_logger.info(
        f"[DEBUG_HEATMAP_FINAL_COLS] heat_ALL.parquetに保存される最終列: {pivot_to_excel_all.columns.tolist()}"
    )

    fp_all_path = out_dir_path / "heat_ALL.parquet"
    try:
        pivot_to_excel_all.to_parquet(fp_all_path)
        log.info(
            "[heatmap.build_heatmap] 全体ヒートマップ (heat_ALL.parquet) 作成完了。"
        )
    except Exception as e_write_all:
        log.error(
            f"[heatmap.build_heatmap] heat_ALL.parquet 作成エラー: {e_write_all}",
            exc_info=True,
        )

    fp_all_xlsx_path = out_dir_path / "heat_ALL.xlsx"
    try:
        save_df_xlsx(pivot_to_excel_all, fp_all_xlsx_path, sheet_name="heat_ALL")
    except Exception as e_xlsx_all:
        log.error(
            f"[heatmap.build_heatmap] heat_ALL.xlsx 作成エラー: {e_xlsx_all}",
            exc_info=True,
        )

    try:
        log.info(f"{fp_all_xlsx_path.name} に書式を設定します。")
        wb = openpyxl.load_workbook(fp_all_xlsx_path)
        ws = wb.active

        data_columns = pivot_to_excel_all.columns.drop(SUMMARY5, errors="ignore")

        _apply_conditional_formatting_to_worksheet(ws, data_columns)
        _apply_holiday_column_styling(ws, data_columns, holidays_set, _parse_as_date)

        wb.save(fp_all_xlsx_path)
        log.info(f"書式設定を {fp_all_xlsx_path.name} に保存しました。")
    except Exception as e:
        log.error(f"{fp_all_xlsx_path.name} への書式設定中にエラー: {e}", exc_info=True)

    unique_roles_list_final_loop = sorted(
        list(set(df_for_heatmap_actuals[role_col_name]))
    )
    log.info(
        f"[heatmap.build_heatmap] 職種別ヒートマップ作成開始。対象: {unique_roles_list_final_loop}"
    )
    for role_item_final_loop in unique_roles_list_final_loop:
        role_safe_name_final_loop = safe_sheet(str(role_item_final_loop))
        log.debug(f"職種 '{role_item_final_loop}' 開始...")
        df_role_subset = df_for_heatmap_actuals[
            df_for_heatmap_actuals[role_col_name] == role_item_final_loop
        ]
        pivot_data_role_actual = pd.DataFrame(index=time_index_labels)
        if not df_role_subset.empty:
            pivot_data_role_actual = (
                df_role_subset.drop_duplicates(
                    subset=["date_lbl", "time", staff_col_name]
                )
                .pivot_table(
                    index="time",
                    columns="date_lbl",
                    values=staff_col_name,
                    aggfunc="nunique",
                    fill_value=0,
                )
                .reindex(index=time_index_labels, fill_value=0)
            )
        pivot_data_role_final = pivot_data_role_actual.reindex(
            columns=all_date_labels_in_period_str, fill_value=0
        )
        if not pivot_data_role_final.columns.empty:
            try:
                cols_to_sort_r = pd.Series(pivot_data_role_final.columns).astype(str)
                valid_date_cols_r = cols_to_sort_r[
                    cols_to_sort_r.str.match(r"^\d{4}-\d{2}-\d{2}$")
                ]
                if not valid_date_cols_r.empty:
                    sorted_cols_r = sorted(
                        valid_date_cols_r,
                        key=lambda d: dt.datetime.strptime(d, "%Y-%m-%d").date(),
                    )
                    other_cols_r = [
                        c
                        for c in pivot_data_role_final.columns
                        if c not in valid_date_cols_r.tolist()
                    ]
                    pivot_data_role_final = pivot_data_role_final[
                        sorted_cols_r + other_cols_r
                    ]
            except Exception as e_sort_r:
                log.warning(f"職種 '{role_item_final_loop}' 日付ソート失敗: {e_sort_r}")

        actual_staff_for_role_need_input = pivot_data_role_actual.copy()
        if not actual_staff_for_role_need_input.empty:
            new_column_map_for_role_need = {}
            for col_str_role in actual_staff_for_role_need_input.columns:
                dt_obj_role = _parse_as_date(str(col_str_role))
                if dt_obj_role:
                    new_column_map_for_role_need[col_str_role] = dt_obj_role

            if new_column_map_for_role_need:
                valid_keys_for_role_rename = [
                    k
                    for k in new_column_map_for_role_need.keys()
                    if k in actual_staff_for_role_need_input.columns
                ]
                actual_staff_for_role_need_input = actual_staff_for_role_need_input[
                    valid_keys_for_role_rename
                ].rename(columns=new_column_map_for_role_need)
            else:
                actual_staff_for_role_need_input = pd.DataFrame(index=time_index_labels)

        # 重要な修正：職種別でも全期間の日付を補完
        if include_zero_days and all_dates_in_period_list:
            for date in all_dates_in_period_list:
                if ref_start_date_for_need <= date <= ref_end_date_for_need and date not in final_holidays_to_use:
                    if date not in actual_staff_for_role_need_input.columns:
                        actual_staff_for_role_need_input[date] = 0

        dow_need_pattern_role_df = calculate_pattern_based_need(
            actual_staff_for_role_need_input,
            ref_start_date_for_need,
            ref_end_date_for_need,
            final_statistic_method,
            need_remove_outliers,
            need_iqr_multiplier,
            slot_minutes_for_empty=slot_minutes,
            holidays=final_holidays_to_use,
            adjustment_factor=need_adjustment_factor,
            include_zero_days=include_zero_days,
            all_dates_in_period=all_dates_in_period_list,
        )

        need_df_role_final = pd.DataFrame(index=time_index_labels, columns=pivot_data_role_final.columns, dtype=float).fillna(0)
        for date_str_col_map in pivot_data_role_final.columns:
            current_date_obj_map = dt.datetime.strptime(date_str_col_map, "%Y-%m-%d").date()
            if current_date_obj_map in holidays_set:
                need_df_role_final[date_str_col_map] = 0
            else:
                day_of_week_map = current_date_obj_map.weekday()
                if day_of_week_map in dow_need_pattern_role_df.columns:
                    need_df_role_final[date_str_col_map] = dow_need_pattern_role_df[day_of_week_map]
                else:
                    need_df_role_final[date_str_col_map] = 0

        need_r_series = need_df_role_final.mean(axis=1).round()

        if upper_calc_method == _("下限値(Need) + 固定値"):
            fixed_val = (upper_calc_param or {}).get("fixed_value", 0)
            upper_r_series = need_r_series + fixed_val
        elif upper_calc_method == _("下限値(Need) * 固定係数"):
            factor = (upper_calc_param or {}).get("factor", 1.0)
            upper_r_series = (need_r_series * factor).apply(np.ceil)
        elif upper_calc_method == _("過去実績のパーセンタイル"):
            pct = (upper_calc_param or {}).get("percentile", 90) / 100
            upper_r_series = pivot_data_role_actual.quantile(pct, axis=1).round()
        else:
            upper_r_series = (
                derive_max_staff(pivot_data_role_actual, max_method)
                if not pivot_data_role_actual.empty
                else pd.Series(0, index=time_index_labels)
            )

        upper_r_series = np.maximum(upper_r_series, need_r_series)
        staff_r_series = (
            pivot_data_role_final.drop(columns=SUMMARY5, errors="ignore")
            .sum(axis=1)
            .round()
        )
        lack_r_series = (need_r_series - staff_r_series).clip(lower=0)
        excess_r_series = (staff_r_series - upper_r_series).clip(lower=0)

        pivot_to_excel_role = pivot_data_role_final.copy()
        for col, data in zip(
            SUMMARY5,
            [
                need_r_series,
                upper_r_series,
                staff_r_series,
                lack_r_series,
                excess_r_series,
            ],
            strict=True,
        ):
            pivot_to_excel_role[col] = data

        fp_role = out_dir_path / f"heat_{role_safe_name_final_loop}.parquet"
        try:
            pivot_to_excel_role.to_parquet(fp_role)
            log.info(f"職種 '{role_item_final_loop}' ヒートマップ作成完了。")
        except Exception as e_role_write:
            log.error(
                f"heat_{role_safe_name_final_loop}.parquet 作成エラー: {e_role_write}",
                exc_info=True,
            )

        fp_role_xlsx = out_dir_path / f"heat_{role_safe_name_final_loop}.xlsx"
        try:
            save_df_xlsx(
                pivot_to_excel_role,
                fp_role_xlsx,
                sheet_name=f"heat_{role_safe_name_final_loop}",
            )
        except Exception as e_role_xlsx:
            log.error(
                f"heat_{role_safe_name_final_loop}.xlsx 作成エラー: {e_role_xlsx}",
                exc_info=True,
            )

        try:
            log.info(f"{fp_role_xlsx.name} に書式を設定します。")
            wb = openpyxl.load_workbook(fp_role_xlsx)
            ws = wb.active
            data_columns = pivot_to_excel_role.columns.drop(SUMMARY5, errors="ignore")
            _apply_conditional_formatting_to_worksheet(ws, data_columns)
            _apply_holiday_column_styling(
                ws, data_columns, holidays_set, _parse_as_date
            )
            wb.save(fp_role_xlsx)
        except Exception as e:
            log.error(f"{fp_role_xlsx.name} への書式設定中にエラー: {e}", exc_info=True)

    # ── Employment heatmaps ───────────────────────────────────────────────
    employment_col_name = "employment"
    unique_employments_list_final_loop = (
        sorted(list(set(df_for_heatmap_actuals[employment_col_name])))
        if employment_col_name in df_for_heatmap_actuals.columns
        else []
    )
    log.info(
        f"[heatmap.build_heatmap] 雇用形態別ヒートマップ作成開始。対象: {unique_employments_list_final_loop}"
    )
    for emp_item_final_loop in unique_employments_list_final_loop:
        emp_safe_name_final_loop = safe_sheet(str(emp_item_final_loop))
        log.debug(f"雇用形態 '{emp_item_final_loop}' 開始...")
        df_emp_subset = df_for_heatmap_actuals[
            df_for_heatmap_actuals[employment_col_name] == emp_item_final_loop
        ]
        pivot_data_emp_actual = pd.DataFrame(index=time_index_labels)
        if not df_emp_subset.empty:
            pivot_data_emp_actual = (
                df_emp_subset.drop_duplicates(
                    subset=["date_lbl", "time", staff_col_name]
                )
                .pivot_table(
                    index="time",
                    columns="date_lbl",
                    values=staff_col_name,
                    aggfunc="nunique",
                    fill_value=0,
                )
                .reindex(index=time_index_labels, fill_value=0)
            )
        pivot_data_emp_final = pivot_data_emp_actual.reindex(
            columns=all_date_labels_in_period_str, fill_value=0
        )
        if not pivot_data_emp_final.columns.empty:
            try:
                cols_to_sort_e = pd.Series(pivot_data_emp_final.columns).astype(str)
                valid_date_cols_e = cols_to_sort_e[
                    cols_to_sort_e.str.match(r"^\d{4}-\d{2}-\d{2}$")
                ]
                if not valid_date_cols_e.empty:
                    sorted_cols_e = sorted(
                        valid_date_cols_e,
                        key=lambda d: dt.datetime.strptime(d, "%Y-%m-%d").date(),
                    )
                    other_cols_e = [
                        c
                        for c in pivot_data_emp_final.columns
                        if c not in valid_date_cols_e.tolist()
                    ]
                    pivot_data_emp_final = pivot_data_emp_final[
                        sorted_cols_e + other_cols_e
                    ]
            except Exception as e_sort_e:
                log.warning(
                    f"雇用形態 '{emp_item_final_loop}' 日付ソート失敗: {e_sort_e}"
                )

        actual_staff_for_emp_need_input = pivot_data_emp_actual.copy()
        if not actual_staff_for_emp_need_input.empty:
            new_column_map_for_emp_need = {}
            for col_str_emp in actual_staff_for_emp_need_input.columns:
                dt_obj_emp = _parse_as_date(str(col_str_emp))
                if dt_obj_emp:
                    new_column_map_for_emp_need[col_str_emp] = dt_obj_emp

            if new_column_map_for_emp_need:
                valid_keys_for_emp_rename = [
                    k
                    for k in new_column_map_for_emp_need.keys()
                    if k in actual_staff_for_emp_need_input.columns
                ]
                actual_staff_for_emp_need_input = actual_staff_for_emp_need_input[
                    valid_keys_for_emp_rename
                ].rename(columns=new_column_map_for_emp_need)
            else:
                actual_staff_for_emp_need_input = pd.DataFrame(index=time_index_labels)

        # 重要な修正：雇用形態別でも全期間の日付を補完
        if include_zero_days and all_dates_in_period_list:
            for date in all_dates_in_period_list:
                if ref_start_date_for_need <= date <= ref_end_date_for_need and date not in final_holidays_to_use:
                    if date not in actual_staff_for_emp_need_input.columns:
                        actual_staff_for_emp_need_input[date] = 0

        dow_need_pattern_emp_df = calculate_pattern_based_need(
            actual_staff_for_emp_need_input,
            ref_start_date_for_need,
            ref_end_date_for_need,
            final_statistic_method,
            need_remove_outliers,
            need_iqr_multiplier,
            slot_minutes_for_empty=slot_minutes,
            holidays=final_holidays_to_use,
            adjustment_factor=need_adjustment_factor,
            include_zero_days=include_zero_days,
            all_dates_in_period=all_dates_in_period_list,
        )

        need_df_emp_final = pd.DataFrame(index=time_index_labels, columns=pivot_data_emp_final.columns, dtype=float).fillna(0)
        for date_str_col_map in pivot_data_emp_final.columns:
            current_date_obj_map = dt.datetime.strptime(date_str_col_map, "%Y-%m-%d").date()
            if current_date_obj_map in holidays_set:
                need_df_emp_final[date_str_col_map] = 0
            else:
                day_of_week_map = current_date_obj_map.weekday()
                if day_of_week_map in dow_need_pattern_emp_df.columns:
                    need_df_emp_final[date_str_col_map] = dow_need_pattern_emp_df[day_of_week_map]
                else:
                    need_df_emp_final[date_str_col_map] = 0

        need_e_series = need_df_emp_final.mean(axis=1).round()
        upper_e_series = (
            derive_max_staff(pivot_data_emp_actual, max_method)
            if not pivot_data_emp_actual.empty
            else pd.Series(0, index=time_index_labels)
        )
        staff_e_series = (
            pivot_data_emp_final.drop(columns=SUMMARY5, errors="ignore")
            .sum(axis=1)
            .round()
        )
        lack_e_series = (need_e_series - staff_e_series).clip(lower=0)
        excess_e_series = (staff_e_series - upper_e_series).clip(lower=0)

        pivot_to_excel_emp = pivot_data_emp_final.copy()
        for col, data in zip(
            SUMMARY5,
            [
                need_e_series,
                upper_e_series,
                staff_e_series,
                lack_e_series,
                excess_e_series,
            ],
            strict=True,
        ):
            pivot_to_excel_emp[col] = data

        fp_emp = out_dir_path / f"heat_emp_{emp_safe_name_final_loop}.parquet"
        try:
            pivot_to_excel_emp.to_parquet(fp_emp)
            log.info(f"雇用形態 '{emp_item_final_loop}' ヒートマップ作成完了。")
        except Exception as e_emp_write:
            log.error(
                f"heat_emp_{emp_safe_name_final_loop}.parquet 作成エラー: {e_emp_write}",
                exc_info=True,
            )

        fp_emp_xlsx = out_dir_path / f"heat_emp_{emp_safe_name_final_loop}.xlsx"
        try:
            save_df_xlsx(
                pivot_to_excel_emp,
                fp_emp_xlsx,
                sheet_name=f"heat_emp_{emp_safe_name_final_loop}",
            )
        except Exception as e_emp_xlsx:
            log.error(
                f"heat_emp_{emp_safe_name_final_loop}.xlsx 作成エラー: {e_emp_xlsx}",
                exc_info=True,
            )

        try:
            log.info(f"{fp_emp_xlsx.name} に書式を設定します。")
            wb = openpyxl.load_workbook(fp_emp_xlsx)
            ws = wb.active
            data_columns = pivot_to_excel_emp.columns.drop(SUMMARY5, errors="ignore")
            _apply_conditional_formatting_to_worksheet(ws, data_columns)
            _apply_holiday_column_styling(
                ws, data_columns, holidays_set, _parse_as_date
            )
            wb.save(fp_emp_xlsx)
        except Exception as e:
            log.error(f"{fp_emp_xlsx.name} への書式設定中にエラー: {e}", exc_info=True)

    all_unique_roles_from_orig_long_df_meta = (
        sorted(list(set(long_df["role"]))) if "role" in long_df.columns else []
    )
    all_unique_employments_from_orig_long_df_meta = (
        sorted(list(set(long_df["employment"])))
        if "employment" in long_df.columns
        else []
    )
    dow_need_pattern_output = (
        overall_dow_need_pattern_df.reset_index()
        .rename(columns={"time": "time"})
        .to_dict(orient="records")
        if not overall_dow_need_pattern_df.empty
        else []
    )  #  index名変更

    write_meta(
        out_dir_path / "heatmap.meta.json",
        slot=slot_minutes,
        roles=all_unique_roles_from_orig_long_df_meta,
        dates=all_date_labels_in_period_str,
        summary_columns=SUMMARY5,
        estimated_holidays=[d.isoformat() for d in sorted(list(holidays or set()))],
        employments=all_unique_employments_from_orig_long_df_meta,
        dow_need_pattern=dow_need_pattern_output,
        need_calculation_params={
            "ref_start_date": ref_start_date_for_need.isoformat(),
            "ref_end_date": ref_end_date_for_need.isoformat(),
            "statistic_method": final_statistic_method,
            "remove_outliers": need_remove_outliers,
            "iqr_multiplier": need_iqr_multiplier if need_remove_outliers else None,
        },
        leave_statistics=leave_stats,  # 休暇統計をメタデータに追加
    )
    validate_need_calculation(need_all_final_for_summary, pivot_data_all_final)
    log.info("[heatmap.build_heatmap] ヒートマップ生成処理完了。")
