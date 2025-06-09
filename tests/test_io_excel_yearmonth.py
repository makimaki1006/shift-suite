import datetime as dt
from pathlib import Path

import pytest
from openpyxl import Workbook
import pandas as pd

from shift_suite.tasks.io_excel import ingest_excel


def _create_excel(path: Path, ym_cell_value: object | None = "2023年10月") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if ym_cell_value is not None:
        ws["A1"] = ym_cell_value
    ws.append(["氏名", "職種", "1", "2"])
    ws.append(["山田", "看護師", "A", "A"])
    wt = wb.create_sheet("勤務区分")
    wt.append(["勤務記号", "開始", "終了"])
    wt.append(["A", "09:00", "18:00"])
    wb.save(path)


def test_ingest_excel_with_year_month(tmp_path: Path) -> None:
    fp = tmp_path / "shift.xlsx"
    _create_excel(fp)
    long_df, _, unknown = ingest_excel(
        fp,
        shift_sheets=["Sheet1"],
        header_row=2,
        slot_minutes=30,
        year_month_cell_location="A1",
    )
    assert not unknown
    dates = sorted({d.date() for d in long_df["ds"]})
    assert dates == [dt.date(2023, 10, 1), dt.date(2023, 10, 2)]


def test_ingest_excel_year_month_missing(tmp_path: Path) -> None:
    fp = tmp_path / "shift.xlsx"
    _create_excel(fp, ym_cell_value=None)
    with pytest.raises(ValueError):
        ingest_excel(
            fp,
            shift_sheets=["Sheet1"],
            header_row=2,
            slot_minutes=30,
            year_month_cell_location="A1",
        )


def test_ingest_excel_with_date_in_year_month_cell(tmp_path: Path) -> None:
    fp = tmp_path / "shift.xlsx"
    _create_excel(fp, ym_cell_value=dt.date(2025, 2, 1))
    long_df, _, unknown = ingest_excel(
        fp,
        shift_sheets=["Sheet1"],
        header_row=2,
        slot_minutes=30,
        year_month_cell_location="A1",
    )
    assert not unknown
    dates = sorted({d.date() for d in long_df["ds"]})
    assert dates == [dt.date(2025, 2, 1), dt.date(2025, 2, 2)]


def test_ingest_excel_leave_via_remarks(tmp_path: Path) -> None:
    fp = tmp_path / "shift.xlsx"

    wt = pd.DataFrame(
        {"勤務記号": ["特A"], "開始": ["09:00"], "終了": ["18:00"], "備考": ["希望休"]}
    )
    df = pd.DataFrame(
        {"氏名": ["山田"], "職種": ["看護師"], "2024-01-01": ["特A"]}
    )
    with pd.ExcelWriter(fp) as writer:
        wt.to_excel(writer, sheet_name="勤務区分", index=False)
        df.to_excel(writer, sheet_name="Sheet1", index=False)

    long_df, wt_df, unknown = ingest_excel(fp, shift_sheets=["Sheet1"], header_row=1)

    assert not unknown
    wt_row = wt_df[wt_df["code"] == "特A"].iloc[0]
    assert wt_row["holiday_type"] == "希望休"
    assert wt_row["parsed_slots_count"] == 0

    assert set(long_df["holiday_type"]) == {"希望休"}
    assert set(long_df["parsed_slots_count"]) == {0}
