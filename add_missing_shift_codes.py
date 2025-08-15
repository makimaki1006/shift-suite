import pandas as pd
from openpyxl import load_workbook

def add_missing_shift_codes():
    """Add missing shift codes to the 勤務区分 sheet"""
    
    excel_path = 'テストデータ_勤務表　勤務時間_トライアル.xlsx'
    
    # Missing codes that need to be added
    missing_codes = [
        {'記号': '早', '開始': '07:00', '終了': '15:00', '備考': '早番'},
        {'記号': '遅', '開始': '13:00', '終了': '21:00', '備考': '遅番'},  
        {'記号': '日', '開始': '09:00', '終了': '17:00', '備考': '日勤'},
        {'記号': '週', '開始': '', '終了': '', '備考': '週休'},
        {'記号': '土', '開始': '', '終了': '', '備考': '土曜休'},
        {'記号': '準', '開始': '16:00', '終了': '00:30', '備考': '準夜勤'},
        {'記号': '白', '開始': '', '終了': '', '備考': '白番'},
    ]
    
    print(f"Loading Excel file: {excel_path}")
    
    # Read the current 勤務区分 sheet
    df_worktime = pd.read_excel(excel_path, sheet_name='勤務区分', dtype=str).fillna('')
    
    print(f"Current 勤務区分 sheet shape: {df_worktime.shape}")
    print(f"Columns: {df_worktime.columns.tolist()}")
    
    # Add missing codes
    for code_info in missing_codes:
        # Check if code already exists
        if code_info['記号'] not in df_worktime['記号'].values:
            print(f"Adding missing code: {code_info['記号']}")
            
            # Create new row
            new_row = pd.Series(code_info, name=len(df_worktime))
            df_worktime = pd.concat([df_worktime, new_row.to_frame().T], ignore_index=True)
        else:
            print(f"Code {code_info['記号']} already exists, skipping")
    
    print(f"Updated 勤務区分 sheet shape: {df_worktime.shape}")
    
    # Load the workbook and update the sheet
    wb = load_workbook(excel_path)
    
    # Remove existing sheet and add updated one
    if '勤務区分' in wb.sheetnames:
        del wb['勤務区分']
    
    # Create new worksheet
    ws = wb.create_sheet('勤務区分')
    
    # Write headers
    for col, header in enumerate(df_worktime.columns, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row, (_, data) in enumerate(df_worktime.iterrows(), 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Save the workbook
    wb.save(excel_path)
    wb.close()
    
    print(f"Successfully updated {excel_path} with missing shift codes")

if __name__ == "__main__":
    add_missing_shift_codes()