#!/usr/bin/env python3
"""
shift_suite機能を使った0:00重複問題の調査
pandasを直接使わずに既存モジュールを活用
"""

import os
import sys
import json
from pathlib import Path

# shift_suiteモジュールをインポート
sys.path.append('/mnt/c/Users/fuji1/OneDrive/デスクトップ/シフト分析')

try:
    import pandas as pd
    from shift_suite.tasks.heatmap import create_heat_all
    from shift_suite.tasks.utils import log
    PANDAS_AVAILABLE = True
except ImportError as e:
    print(f"pandas利用不可: {e}")
    PANDAS_AVAILABLE = False

print('=== 0:00重複問題の簡易調査 ===')

# 分析対象フォルダ
analysis_folder = '/mnt/c/Users/fuji1/OneDrive/デスクトップ/シフト分析/temp_analysis_results/out_p25_based/'

def check_file_structure():
    """ファイル構造の確認"""
    print('\n1. ファイル構造の確認')
    print('=' * 50)
    
    key_files = [
        'heat_ALL.parquet',
        'heat_ALL.xlsx', 
        'work_patterns.parquet',
        'intermediate_data.parquet',
        'heatmap.meta.json',
        'shortage.meta.json'
    ]
    
    for file in key_files:
        file_path = os.path.join(analysis_folder, file)
        if os.path.exists(file_path):
            size = os.path.getsize(file_path)
            print(f'  ✓ {file} ({size:,} bytes)')
        else:
            print(f'  ✗ {file} (見つかりません)')

def check_meta_files():
    """メタデータファイルの確認"""
    print('\n2. メタデータファイルの確認')
    print('=' * 50)
    
    meta_files = ['heatmap.meta.json', 'shortage.meta.json']
    
    for meta_file in meta_files:
        meta_path = os.path.join(analysis_folder, meta_file)
        if os.path.exists(meta_path):
            print(f'\n{meta_file}の内容:')
            try:
                with open(meta_path, 'r', encoding='utf-8') as f:
                    meta_data = json.load(f)
                    
                    # 重要な情報を抽出
                    if 'time_slots' in meta_data:
                        print(f'  時間スロット数: {meta_data["time_slots"]}')
                    if 'date_range' in meta_data:
                        print(f'  日付範囲: {meta_data["date_range"]}')
                    if 'total_staff' in meta_data:
                        print(f'  総スタッフ数: {meta_data["total_staff"]}')
                    if 'roles' in meta_data:
                        print(f'  役職数: {len(meta_data["roles"])}')
                    
                    # 0:00関連の情報
                    if 'midnight_analysis' in meta_data:
                        print(f'  0:00分析データ: {meta_data["midnight_analysis"]}')
                    
            except Exception as e:
                print(f'  エラー: {e}')
        else:
            print(f'{meta_file}: 見つかりません')

def analyze_excel_heat_all():
    """heat_ALL.xlsxファイルの分析"""
    print('\n3. heat_ALL.xlsxファイルの分析')
    print('=' * 50)
    
    excel_path = os.path.join(analysis_folder, 'heat_ALL.xlsx')
    if not os.path.exists(excel_path):
        print('heat_ALL.xlsxが見つかりません')
        return
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f'シート名: {wb.sheetnames}')
        
        # メインシートの分析
        if 'Heat' in wb.sheetnames:
            ws = wb['Heat']
            print(f'Heat シートの範囲: {ws.calculate_dimension()}')
            
            # 最初の数行を確認
            print('\n先頭データ:')
            for row in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(6, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                print(f'  行{row}: {", ".join(row_data)}')
            
            # 0:00(通常1列目)の値を確認
            print('\n0:00時刻の人員数(推定):')
            time_col = 2  # 通常時刻は2列目から
            for row in range(2, min(12, ws.max_row + 1)):
                date_val = ws.cell(row=row, column=1).value
                staff_val = ws.cell(row=row, column=time_col).value
                if date_val and staff_val is not None:
                    print(f'  {date_val}: {staff_val}人')
        
    except ImportError:
        print('openpyxlが利用できません')
    except Exception as e:
        print(f'Excel分析エラー: {e}')

def check_shortage_files():
    """shortage関連ファイルの確認"""
    print('\n4. shortage関連ファイルの確認')
    print('=' * 50)
    
    shortage_files = [f for f in os.listdir(analysis_folder) if f.startswith('shortage_')]
    print(f'shortage関連ファイル数: {len(shortage_files)}')
    
    for file in shortage_files[:5]:  # 最初の5ファイル
        file_path = os.path.join(analysis_folder, file)
        size = os.path.getsize(file_path)
        print(f'  {file}: {size:,} bytes')

def analyze_with_pandas():
    """pandasが利用可能な場合の詳細分析"""
    if not PANDAS_AVAILABLE:
        print('\n5. pandas詳細分析: 利用不可')
        return
    
    print('\n5. pandas詳細分析')
    print('=' * 50)
    
    try:
        # heat_ALL.parquetの読み込み
        heat_all_path = os.path.join(analysis_folder, 'heat_ALL.parquet')
        if os.path.exists(heat_all_path):
            df = pd.read_parquet(heat_all_path)
            print(f'heat_ALL.parquet: {df.shape}')
            print(f'列: {df.columns.tolist()}')
            
            # 0:00の分析
            if 'slot' in df.columns and 'staff' in df.columns:
                midnight_data = df[df['slot'] == 0]
                if len(midnight_data) > 0:
                    print(f'\n0:00データ行数: {len(midnight_data)}')
                    print(f'0:00平均人員: {midnight_data["staff"].mean():.1f}')
                    print(f'0:00最大人員: {midnight_data["staff"].max()}')
                    print(f'0:00最小人員: {midnight_data["staff"].min()}')
                    
                    # 他の時間との比較
                    other_slots = df[df['slot'].isin([47, 1])]  # 23:30, 0:30
                    if len(other_slots) > 0:
                        other_mean = other_slots.groupby('slot')['staff'].mean()
                        print(f'\n比較時間帯の平均人員:')
                        for slot, mean_val in other_mean.items():
                            time_str = f'{slot//2:02d}:{(slot%2)*30:02d}'
                            print(f'  Slot {slot} ({time_str}): {mean_val:.1f}')
                        
                        midnight_mean = midnight_data['staff'].mean()
                        other_overall_mean = other_slots['staff'].mean()
                        
                        print(f'\n重複問題の定量化:')
                        print(f'  0:00平均: {midnight_mean:.1f}人')
                        print(f'  他時間平均: {other_overall_mean:.1f}人')
                        print(f'  差分: {midnight_mean - other_overall_mean:+.1f}人')
                        print(f'  比率: {midnight_mean / other_overall_mean:.2f}倍')
                        
                        if midnight_mean > other_overall_mean * 1.2:
                            excess = midnight_mean - other_overall_mean
                            print(f'  ⚠️ 0:00で約{excess:.1f}人の重複が疑われます!')
    
    except Exception as e:
        print(f'pandas分析エラー: {e}')

def summary_findings():
    """調査結果の要約"""
    print('\n6. 調査結果の要約')
    print('=' * 50)
    
    findings = []
    
    # ファイル存在確認
    key_files = ['heat_ALL.parquet', 'heat_ALL.xlsx', 'work_patterns.parquet']
    missing_files = []
    for file in key_files:
        if not os.path.exists(os.path.join(analysis_folder, file)):
            missing_files.append(file)
    
    if missing_files:
        findings.append(f'❌ 重要ファイル不足: {", ".join(missing_files)}')
    else:
        findings.append('✅ 主要分析ファイルは存在')
    
    # メタデータ確認
    meta_path = os.path.join(analysis_folder, 'heatmap.meta.json')
    if os.path.exists(meta_path):
        findings.append('✅ ヒートマップメタデータ存在')
    else:
        findings.append('❌ ヒートマップメタデータ不足')
    
    print('主な発見事項:')
    for i, finding in enumerate(findings, 1):
        print(f'  {i}. {finding}')

# メイン実行
if __name__ == '__main__':
    # 分析対象フォルダの確認
    if not os.path.exists(analysis_folder):
        print(f'エラー: {analysis_folder} が見つかりません')
        exit(1)
    
    print(f'分析対象: {analysis_folder}')
    
    # 各分析の実行
    check_file_structure()
    check_meta_files()
    analyze_excel_heat_all()
    check_shortage_files()
    analyze_with_pandas()
    summary_findings()
    
    print('\n=== 簡易調査完了 ===')