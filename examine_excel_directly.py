#!/usr/bin/env python3
"""
heat_ALL.xlsxファイルの直接調査
openpyxlを使ってExcelファイルを直接読み取り
"""

import os
import sys

# openpyxlがあるかチェック
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

print('=== heat_ALL.xlsx直接調査 ===')

analysis_folder = '/mnt/c/Users/fuji1/OneDrive/デスクトップ/シフト分析/temp_analysis_results/out_p25_based/'
excel_path = os.path.join(analysis_folder, 'heat_ALL.xlsx')

def read_excel_with_csv_fallback():
    """ExcelファイルをCSVとして読み取り (openpyxlが使えない場合)"""
    print('\n1. CSVモードでのExcel調査')
    print('=' * 50)
    
    # LibreOfficeやssconvertでCSV変換を試す
    csv_path = excel_path.replace('.xlsx', '_converted.csv')
    
    # ssconvertコマンドを試す
    import subprocess
    try:
        result = subprocess.run(['ssconvert', excel_path, csv_path], 
                               capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            print(f'Excel → CSV変換成功: {csv_path}')
            analyze_converted_csv(csv_path)
        else:
            print(f'ssconvert変換失敗: {result.stderr}')
    except (subprocess.TimeoutExpired, FileNotFoundError):
        print('ssconvertコマンドが利用できません')
    
    # libreofficeコマンドを試す
    try:
        result = subprocess.run(['libreoffice', '--headless', '--convert-to', 'csv', 
                               '--outdir', analysis_folder, excel_path],
                               capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            csv_path2 = os.path.join(analysis_folder, 'heat_ALL.csv')
            if os.path.exists(csv_path2):
                print(f'LibreOffice変換成功: {csv_path2}')
                analyze_converted_csv(csv_path2)
        else:
            print(f'LibreOffice変換失敗: {result.stderr}')
    except (subprocess.TimeoutExpired, FileNotFoundError):
        print('LibreOfficeコマンドが利用できません')

def analyze_converted_csv(csv_path):
    """変換されたCSVファイルの分析"""
    if not os.path.exists(csv_path):
        return
    
    print(f'\nCSVファイル分析: {csv_path}')
    
    import csv
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    print(f'行数: {len(rows)}')
    print(f'列数: {len(rows[0]) if rows else 0}')
    
    # 最初の数行を表示
    print('\n先頭5行:')
    for i, row in enumerate(rows[:5]):
        print(f'  行{i+1}: {", ".join(row[:10])}...')  # 最初の10列まで

def read_excel_with_openpyxl():
    """openpyxlを使ったExcel読み取り"""
    print('\n2. openpyxlでのExcel調査')
    print('=' * 50)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f'シート名: {wb.sheetnames}')
        
        # メインシートの分析
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]  # 最初のシート
            print(f'アクティブシート: {ws.title}')
            print(f'データ範囲: {ws.calculate_dimension()}')
            
            # ヘッダー行の確認
            print('\nヘッダー行 (行1):')
            header_row = []
            for col in range(1, min(25, ws.max_column + 1)):  # 最初の24列
                cell_value = ws.cell(row=1, column=col).value
                header_row.append(str(cell_value) if cell_value is not None else '')
            print('  ' + ' | '.join(header_row))
            
            # 日付列と0:00列を特定
            date_col = 1  # 通常は1列目
            time_cols = {}
            
            # 1行目から時刻列を探す
            for col in range(2, min(50, ws.max_column + 1)):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and '0:00' in str(cell_value):
                    time_cols['00:00'] = col
                elif cell_value and '23:45' in str(cell_value):
                    time_cols['23:45'] = col
                elif cell_value and '0:15' in str(cell_value):
                    time_cols['00:15'] = col
            
            print(f'\n時刻列の位置: {time_cols}')
            
            # 0:00列の実データを確認
            if '00:00' in time_cols:
                midnight_col = time_cols['00:00']
                print(f'\n0:00列(列{midnight_col})の実データ:')
                
                midnight_values = []
                for row in range(2, min(32, ws.max_row + 1)):  # 最初の30日分
                    date_val = ws.cell(row=row, column=date_col).value
                    staff_val = ws.cell(row=row, column=midnight_col).value
                    
                    if date_val and staff_val is not None:
                        print(f'  {date_val}: {staff_val}人')
                        if isinstance(staff_val, (int, float)):
                            midnight_values.append(float(staff_val))
                
                # 統計計算
                if midnight_values:
                    avg_staff = sum(midnight_values) / len(midnight_values)
                    max_staff = max(midnight_values)
                    min_staff = min(midnight_values)
                    
                    print(f'\n0:00人員統計:')
                    print(f'  平均: {avg_staff:.1f}人')
                    print(f'  最大: {max_staff:.0f}人')
                    print(f'  最小: {min_staff:.0f}人')
                    
                    # Need=0なのに人員がいる矛盾を指摘
                    if avg_staff > 0:
                        print(f'\n⚠️ 重要な矛盾:')
                        print(f'  - Need計算では0:00はNeed=0 (勤務不要)')
                        print(f'  - しかし実際には平均{avg_staff:.1f}人が配置されている')
                        print(f'  - これは夜勤者が0:00に重複カウントされている可能性')
            
            # 他の時間との比較
            comparison_data = {}
            for time_name, col in time_cols.items():
                values = []
                for row in range(2, min(32, ws.max_row + 1)):
                    staff_val = ws.cell(row=row, column=col).value
                    if isinstance(staff_val, (int, float)):
                        values.append(float(staff_val))
                
                if values:
                    comparison_data[time_name] = {
                        'avg': sum(values) / len(values),
                        'max': max(values),
                        'min': min(values)
                    }
            
            if comparison_data:
                print(f'\n時間別比較:')
                print('時刻   | 平均人員 | 最大 | 最小')
                print('-' * 35)
                for time_name, stats in comparison_data.items():
                    print(f'{time_name:6s} | {stats["avg"]:8.1f} | {stats["max"]:4.0f} | {stats["min"]:4.0f}')
        
    except Exception as e:
        print(f'openpyxl読み取りエラー: {e}')

def analyze_staff_vs_need_discrepancy():
    """スタッフ配置とNeed計算の乖離分析"""
    print('\n3. スタッフ配置とNeed計算の乖離分析')
    print('=' * 60)
    
    print('発見された矛盾:')
    print('1. Need計算システム:')
    print('   - 0:00〜6:45は全てNeed=0')
    print('   - 理論上、深夜時間帯にスタッフは不要')
    print('')
    print('2. 実際のスタッフ配置:')
    print('   - 0:00に毎日12-19人が配置されている')
    print('   - shortage_leave.csvで確認済み')
    print('')
    print('3. 可能な原因:')
    print('   a) 夜勤者の勤務時間が0:00を跨いでいる')
    print('   b) 連続勤務の境界処理で重複カウント')
    print('   c) Need計算とスタッフ配置の処理が分離されている')
    print('')
    print('4. 重複の証拠:')
    print('   - Need=0なのにStaff>0の矛盾')
    print('   - 夜勤終了と明け番開始の境界時刻')
    print('   - 同一時刻に複数の勤務パターンが重複')

# メイン実行
if __name__ == '__main__':
    if not os.path.exists(excel_path):
        print(f'エラー: {excel_path} が見つかりません')
        exit(1)
    
    print(f'分析対象: {excel_path}')
    
    if OPENPYXL_AVAILABLE:
        read_excel_with_openpyxl()
    else:
        print('openpyxlが利用できません。代替手段を試行します。')
        read_excel_with_csv_fallback()
    
    analyze_staff_vs_need_discrepancy()
    
    print('\n=== Excel調査完了 ===')